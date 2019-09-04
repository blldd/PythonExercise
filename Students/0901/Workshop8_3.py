# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""

import sqlite3
import openpyxl

import matplotlib.pyplot as plt


def get_data(self):
    wb = openpyxl.load_workbook(self.file_name)
    sheet = wb[self.sheet.name]
    test_data = []
    # 读取Excel某张表中所有行，1-4列的数据
    for i in range(1, sheet.max_row + 1):
        sub_data = {}
        sub_data['method'] = sheet.cell(i, 1).value
        sub_data['url'] = sheet.cell(i, 2).value
        sub_data['data'] = sheet.cell(i, 3).value
        sub_data['expected'] = sheet.cell(i, 4).value
        test_data.append(sub_data)

    return test_data


"""
Generate a line graph using matplotlib to compare the yearly average temperatures for 
Sydney,London, New York and Jakarta for the 20th century (1900-2000, one line for each city).

Include the relevant axis titles and data labels using the legend() function. You may need
to read https://matplotlib.org/api/pyplot_api.html#matplotlib.pyplot.legend
"""

wb = openpyxl.load_workbook("ClimateData.xlsx")

sheet = wb["Sheet1"]

x_ = list(range(1900, 2001))
dictt = {"Sydney": [],
         "London": [],
         "New York": [],
         "Jakarta": []}

tmp = 0
cnt = 0
for i in range(2, sheet.max_row + 2):

    if cnt < 12:
        cnt += 1
        temperature = sheet.cell(i, 2).value
        if temperature:
            tmp += temperature
    elif cnt == 12:
        city = sheet.cell(i, 4).value
        dictt[city].append(tmp / 12)

        cnt = 0
        tmp = 0

plt.figure(figsize=(20, 15))

plt.xlabel('Year')
plt.ylabel('yearly average temperatures')
plt.plot(x_, dictt['Sydney'][:101], "or", x_, dictt['Sydney'][:101], "--,b")
plt.plot(x_, dictt['London'][:101], "or", x_, dictt['London'][:101], "--,g")
plt.plot(x_, dictt['New York'][:101], "or", x_, dictt['New York'][:101], "--,r")
plt.plot(x_, dictt['Jakarta'][:101], "or", x_, dictt['Jakarta'][:101], "--,y")
plt.title('yearly average temperatures for Sydney,London, New York and Jakarta for the 20th century')

plt.legend()
plt.savefig('temperatures_data')

plt.show()
